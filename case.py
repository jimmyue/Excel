import openpyxl
import xlsxwriter

class DropDown():
    def write_with_dropdown(self, book_name, sheet_name="Sheet1"):
        #官网https://xlsxwriter.readthedocs.io/ 写入下拉列表
        workbook = xlsxwriter.Workbook(book_name)
        worksheet = workbook.add_worksheet(sheet_name)
        worksheet.data_validation("A1", {'validate':'list', 'source':[1, 2, 3, 4]})
        workbook.close()
 
    def read_with_dropdown(self, book_name, sheet_name="Sheet1"):
        # https://blog.csdn.net/weixin_41267342/article/details/86634007。读取下拉列表
        wb = openpyxl.load_workbook(book_name)  #读取excel
        #读取sheet表
        ws = wb[sheet_name]
        # 读取excel指定单元格数据
        # data = ws["A1":"G5"]
        # 获取内容存在下拉选的框数据
        validations = ws.data_validations.dataValidation
        # 遍历存在下拉选的单元格
        for validation in validations:
            # 获取下拉框中的所有选择值
            cell = validation.sqref
            result = validation.formula1
            print("单元格位置:" + str(cell) + ",下拉选内容：" + result)


if __name__ == "__main__":
	dd = DropDown()
	dd.write_with_dropdown('test.xlsx', "下拉")
	dd.read_with_dropdown('test.xlsx', "下拉")
